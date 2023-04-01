import os

import openpyxl
import fofa_config



class File_with:

    def __init__(self):
        self.xlsx_filename="./结果.xlsx"
        self.txt_filename = "./结果.txt"
        self.check_filename = "./待检测.txt"
        fofa_config.config_init()
        self.file_name_check()
        self. wb2 = openpyxl.Workbook()
        self.flag=True

    #报存文件名函数，没有设置使用默认文件名
    def file_name_check(self):
        if os.path.splitext(fofa_config.file_path)[-1]!= '':
            hz = os.path.splitext(fofa_config.file_path)[-1]
            if hz == '.txt':
                self.txt_filename = fofa_config.file_path
            elif hz == '.xlsx':
                self.xlsx_filename = fofa_config.file_path
            else:
                pass
        else:
            name = os.path.basename(fofa_config.file_path)
            path=os.path.dirname(fofa_config.file_path)
            if name == '':
                self.txt_filename = path + '/结果.txt'
                self.xlsx_filename = path + '/结果.xlsx'
            else:
                self.txt_filename = path + '/'+ name +'.txt'
                self.xlsx_filename = path + '/'+ name +'.xlsx'



    #xlsx文件保存操作
    def file_xlsx(self,file_name = None):
        if file_name == None:
            self.wb2.save(self.xlsx_filename)
        else:
            self.wb2.save(file_name)


    #xlsx文件写入操作
    def file_save_xlsx(self,data,save):
        self.ws = self.wb2.active
        if self.flag:
            self.flag = False
            if save == '极简数据':
                self.ws['A' + str(1)] = data[0]
                self.ws['B' + str(1)] = data[1]
            elif save == '基本数据':
                self.ws['A' + str(1)] = data[0]
                self.ws['B' + str(1)] = data[1]
                self.ws['C' + str(1)] = data[2]
                self.ws['D' + str(1)] = data[3]
            else:
                self.ws['A' + str(1)] = data[0]
                self.ws['B' + str(1)] = data[1]
                self.ws['C' + str(1)] = data[2]
                self.ws['D' + str(1)] = data[3]
                if len(data)>4:
                    self.ws['E' + str(1)] = data[4]
                    self.ws['F' + str(1)] = data[5]
                    self.ws['F' + str(1)] = data[6]

        else:
            if save == '极简数据':
                self.ws.append([data[0], data[1]])
            elif save == '基本数据':
                self.ws.append([data[0], data[1], data[2], data[3]])
            else:
                if len(data) > 4:
                    self.ws.append([data[0], data[1], data[2], data[3], data[4], data[5], data[6]])
                else:
                    self.ws.append([data[0], data[1], data[2], data[3]])


    #txt文件写入操作（此方法只保存host数据）
    def file_save_txt(self,data):
        data[0] = data[0].strip("/")
        if self.flag:
            with open(self.txt_filename, 'w') as f:
                # print('正在创建文件')
                self.flag = False
                f.write(data[0] + "/" + "\n")
        else:
            with open(self.txt_filename, 'a') as f:
                # print("写入文件")
                f.write(data[0] + "/" + "\n")



    #离线验证打开需要验证的目标文件，仅支持.txt文件
    def file_poc_check(self):
        with open(self.check_filename, 'r') as f:
            li_list = []
            for data in f.readlines():
                li = [data.strip()]
                li_list.append(li)

            return li_list



fofa_file=File_with()